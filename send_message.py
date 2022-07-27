from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from helium import *
import time
from openpyxl import load_workbook
from pathlib import Path
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import pandas as pd
import datetime
import string, random
from utils import *
from selenium import webdriver
from random import randint
from time import sleep


df = pd.read_excel('message.xlsx')
name = df['Name'].values.tolist()
number = df['Number'].values.tolist()
message = df['Message'].values.tolist()
pdf = df['Pdf'].values.tolist()
error = {}
timeout_secs=200
browser = start_chrome('https://web.whatsapp.com')
browser.fullscreen_window()
# wait_until(Text('Keep your phone').exists,timeout_secs=timeout_secs )
wait_until(Text('Now send and receive messages without keeping your phone online.').exists,timeout_secs=timeout_secs )

j=0
# for i in number:
#     try:
#         search_box = browser.find_element_by_xpath('//*[@id="side"]/div[1]/div/label/div/div[2]')
#         search_box.clear()
#         time.sleep(2)
#         search_box.send_keys(i)
#         time.sleep(2)
#         k = name[j]
#         try:
#             user = browser.find_element_by_xpath(f'//span[@title="{k}"]')
#             time.sleep(2)
#             user.click()
#         except:
#             user = browser.find_element_by_xpath(f'//span[@title="{i}"]')
#             time.sleep(2)
#             user.click()
#         time.sleep(2)
# #         -------------send message-----------
#         try:
#             message_box = browser.find_element_by_xpath('//*[@id="main"]/footer/div[1]/div[2]/div/div[1]/div/div[2]')
#             message_box.click()
#             time.sleep(2)
#             message2 =  message[j] + f"  Message Id: {''.join(random.choices(string.ascii_letters + string.digits, k=8))}"
#             message1 = message2.split('\\n')
            
#             for msg in message1:
#                 action = ActionChains(browser)
#                 action.send_keys_to_element(message_box, msg)
#                 action.key_down(Keys.CONTROL)
#                 action.key_down(Keys.ENTER)
#                 action.key_up(Keys.ENTER)
#                 action.key_up(Keys.CONTROL)
#                 action.perform()
#             time.sleep(2)
#             send = browser.find_element_by_xpath('//*[@id="main"]/footer/div[1]/div[2]/div/div[2]/button')
#             send.click()
#             time.sleep(2)
#         except:
#             print('no message found!')
#         try:
#     #         ----------send file--------------
#             file_path = str((Path(pdf[j])).absolute())
#             browser.find_element_by_css_selector("span[data-icon='clip']").click()
#             time.sleep(2)
#             browser.find_element_by_css_selector("input[type='file']").send_keys(file_path)
#             time.sleep(3)
#             try:
#     #             for small file
#                 send = browser.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/div/div[2]/div[2]/div/div/span')
#                 send.click()
#                 time.sleep(5)
#             except:
#     #             for large files
#                 send = browser.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/div/div[2]/div[2]/div/div/span')
#                 send.click()
#                 time.sleep(15)
#         except:
#             print('no attachment found to send to contact {}'.format(i))
#     except:
#         error['{}'.format(i)] = 'no contact fount'

#     j = j+1
# dfs = pd.DataFrame(data=error, index=[0])
# dfs = (dfs.T)
# x = datetime.datetime.now()
# y  = x.strftime("%m_%d_%Y_%H_%M_%S")
# dfs.to_excel('error_{}.xlsx'.format(y))


# code for looping

x = datetime.datetime.now()
y  = x.strftime("%m_%d_%Y_%H_%M_%S")
writer = pd.ExcelWriter('error_{}.xlsx'.format(y), engine='xlsxwriter')
op = 'error_{}.xlsx'.format(y)
writer.save()
writer.close()
rdf = read_df(op)
j = 0
while(len(number)!=0):
    try:
        num = int(number[0])
        print('printing number', num)
        search_box = browser.find_element_by_xpath('//*[@id="side"]/div[1]/div/div/div[2]/div/div[2]')
        search_box.clear()
        search_box.click()
        search_box.send_keys(num)
        time.sleep(2)
        k = name[0]


        # search_box = browser.find_element_by_xpath('//*[@id="side"]/div[1]/div/label/div/div[2]')
        # search_box.clear()
        # time.sleep(2)
        # search_box.send_keys(number[0])
        # time.sleep(2)
        # k = name[0]
        try:
            user = browser.find_element_by_xpath(f'//span[@title="{k}"]')
            time.sleep(1.5+0.3*randint(1,5))
            user.click()
        except:
            user = browser.find_element_by_xpath(f'//span[@title="{number[0]}"]')
            time.sleep(1.5+0.3*randint(1,5))
            user.click()
        time.sleep(2)
        #         -------------send message-----------
        try:
            message_box = browser.find_element_by_xpath('//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[1]/div/div[2]')
            message_box.click()
            time.sleep(1+0.3*randint(1,6))
            # to add random id to message keep line no 133 and 134 un commented and comment line no 137
            #message2 =  message[0] + f"  Message Id: {''.join(random.choices(string.ascii_letters + string.digits, k=8))}"
            # message1 = message2.split("\\n")

            # to send message without random id comment line 133 and 134 and uncomment line no 137
            message1 = message[0].split("#")
            # separator can be changed as needed.Currently set to '#'
            br = (Keys.SHIFT) + (Keys.ENTER) + (Keys.SHIFT)
            # print(message1)
            for msg in message1:
                # print(msg)
                message_box.send_keys(msg+br)

                # action = ActionChains(browser)
                # action.send_keys_to_element(message_box, msg)
                # action.key_down(Keys.SHIFT)
                # action.key_down(Keys.ENTER)
                # action.key_up(Keys.ENTER)
                # action.key_up(Keys.SHIFT)
                # action.key_down(Keys.CONTROL)
                # action.key_down(Keys.ENTER)
                # action.key_up(Keys.ENTER)
                # action.key_up(Keys.CONTROL)
                # action.perform()
                # time.sleep(2)
            time.sleep(0.3*randint(1,5))
            send = browser.find_element_by_xpath('//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[2]/button')
            send.click()
            send_time = datetime.datetime.now()
            time.sleep(2)
        except:
            print('no message found!')
        try:
    #         ----------send file--------------
            p = pdf[0].split(';')
            for i in range(0,len(p)):
                file_path = str((Path(p[i])).absolute())      
                browser.find_element_by_css_selector("span[data-icon='clip']").click()
                browser.find_element_by_css_selector("input[type='file']").send_keys(file_path)
                time.sleep(5)
                try:
        #             for small file
                    send = browser.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/div/div[2]/div[2]/div/div/span')
                    send.click()
                    time.sleep(1)
                except:
        #             for large files
                    send = browser.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/div/div[2]/div[2]/div/div/span')
                    send.click()
                    time.sleep(2)
        except:
            print('no attachment found to send to contact {}'.format(0))
        
        error['number'] = number[0]
        error['name'] = name[0]
        error['message'] = message[0]
        error['Status'] = 'sent'
        error['time'] = send_time.strftime("%m/%d/%Y:%H-%M-%S")
        number.remove(number[0])
        name.remove(name[0])
        message.remove(message[0])
        pdf.remove(pdf[0])
        # logging message record
        dfs = pd.DataFrame(data=error, index=[0])
        rdf = pd.concat([rdf,dfs])
        save_df(op, rdf)
        # dfs = (dfs.T)
        # writer = pd.ExcelWriter('error_{}.xlsx'.format(y), engine='openpyxl')
        # writer.book = load_workbook('error_{}.xlsx'.format(y))
        # writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        # reader = pd.read_excel(r'error_{}.xlsx'.format(y))
        # dfs.to_excel(writer,index=True,header=True,startrow=len(reader)+1)
        # writer.close()
        error.clear()
    except:
        j = j+1
        if j == 4:
            error['number'] = number[0]
            error['name'] = name[0]
            error['message'] = message[0]
            error['Status'] = 'Not sent'
            # logging message record
            dfs = pd.DataFrame(data=error, index=[0])
            rdf = pd.concat([rdf,dfs])
            save_df(op, rdf)
            # dfs = (dfs.T)s
            # writer = pd.ExcelWriter('error_{}.xlsx'.format(y), engine='openpyxl')
            # writer.book = load_workbook('error_{}.xlsx'.format(y))
            # writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            # reader = pd.read_excel(r'error_{}.xlsx'.format(y))
            # dfs.to_excel(writer,index=True,header=True,startrow=len(reader)+1)
            # writer.close()
            error.clear()
            number.remove(number[0])
            name.remove(name[0])
            message.remove(message[0])
            pdf.remove(pdf[0])
            j = 0



# dfs = pd.DataFrame(data=error, index=[0])
# dfs = (dfs.T)
# x = datetime.datetime.now()
# y  = x.strftime("%m_%d_%Y_%H_%M_%S")
# dfs.to_excel('error_{}.xlsx'.format(y))
time.sleep(10)
kill_browser()